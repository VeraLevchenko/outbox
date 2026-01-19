from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from app.models.outbox_journal import OutboxJournal


class ExcelService:
    """Сервис для работы с Excel файлами"""

    def generate_journal_xlsx(self, entries: List[OutboxJournal]) -> BytesIO:
        """
        Генерирует XLSX файл с записями журнала

        Args:
            entries: Список записей журнала

        Returns:
            BytesIO объект с Excel файлом
        """
        # Создаем workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Журнал исходящих"

        # Заголовок
        headers = [
            "№ п/п",
            "Исходящий номер",
            "Дата",
            "Кому",
            "Исполнитель",
            "Путь к файлам"
        ]

        # Стили для заголовка
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Устанавливаем ширину колонок
        ws.column_dimensions['A'].width = 8   # № п/п
        ws.column_dimensions['B'].width = 15  # Исходящий номер
        ws.column_dimensions['C'].width = 12  # Дата
        ws.column_dimensions['D'].width = 40  # Кому
        ws.column_dimensions['E'].width = 25  # Исполнитель
        ws.column_dimensions['F'].width = 50  # Путь

        # Стили для данных
        data_alignment = Alignment(vertical="top", wrap_text=True)

        # Записываем данные
        for row_num, entry in enumerate(entries, 2):
            # № п/п
            cell = ws.cell(row=row_num, column=1)
            cell.value = row_num - 1
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            # Исходящий номер
            cell = ws.cell(row=row_num, column=2)
            cell.value = entry.outgoing_no
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            # Дата
            cell = ws.cell(row=row_num, column=3)
            cell.value = entry.outgoing_date.strftime("%d.%m.%Y") if entry.outgoing_date else ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

            # Кому
            cell = ws.cell(row=row_num, column=4)
            cell.value = entry.to_whom or ""
            cell.alignment = data_alignment
            cell.border = border

            # Исполнитель
            cell = ws.cell(row=row_num, column=5)
            cell.value = entry.executor or ""
            cell.alignment = data_alignment
            cell.border = border

            # Путь
            cell = ws.cell(row=row_num, column=6)
            cell.value = entry.folder_path or ""
            cell.alignment = data_alignment
            cell.border = border

        # Закрепляем первую строку
        ws.freeze_panes = "A2"

        # Сохраняем в BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer


# Singleton instance
excel_service = ExcelService()
